from django.shortcuts import render, redirect
from ..utils import get_all_sheets,get_employee_list
from django.conf import settings
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import calendar

MONTHLY_SHEETS = []
EMPLOYEE_LIST = []


def initialize(request):
    initialize_excel()

    global MONTHLY_SHEETS
    global EMPLOYEE_LIST
    MONTHLY_SHEETS = get_all_sheets()
    EMPLOYEE_LIST = get_employee_list()

    return render(
        request,
        settings.TEMPLATE_INDEX,
        {
            "employee_list": EMPLOYEE_LIST,
        },
    )

def initialize_excel():

    ret = False
    wb = openpyxl.load_workbook(settings.EXCEL_PATH)
    try:
        # create sheets for 3 months ahead
        current_date = datetime.now()
        for i in range(3):
            sheet_date = current_date + timedelta(days=i * 30)
            sheet_name = (current_date + timedelta(days=i * 30)).strftime("%Y-%m")

            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(title=sheet_name)
            else:
                sheet = wb[sheet_name]

            sheet.cell(row=2, column=1, value="SeatID")  # Same: sheet["A1"] = "Seat ID"
            first_day, last_day = (
                1,
                calendar.monthrange(sheet_date.year, sheet_date.month)[1],
            )

            dayIdx = 1
            for day in range(first_day, last_day + 1):

                date_obj = datetime(sheet_date.year, sheet_date.month, day)
                date = date_obj.strftime("%m-%d")
                weekday = date_obj.strftime("%a")

                # plan-B: not to show weekends
                if weekday in ('Sat','Sun'): continue
                # plan-B

                col_am = 2 + (dayIdx - 1) * 2
                col_pm = col_am + 1
                sheet.cell(row=1, column=2 + (dayIdx - 1) * 2, value=f"{date}({weekday})")
                sheet.cell(row=2, column=col_am, value=f"AM")
                sheet.cell(row=2, column=col_pm, value=f"PM")
                dayIdx += 1

            sheet_seat = wb[settings.EXCEL_SHEET_NAME_SEAT]

            for idx, row in enumerate(
                sheet_seat.iter_rows(min_row=2, max_col=2, values_only=True)
            ):
                if row[0]:
                    sheet.cell(row=3 + idx, column=1, value=row[0])

            ret = True
    
    except Exception as e:
        print(e)
        ret = False
    finally:
        wb.save(settings.EXCEL_PATH)
        wb.close()
        return ret

def checkSelection(request):
    if request.method == 'POST':
        employeeId = request.POST.get('employeeSelect')

        employeeName = None
        for option in EMPLOYEE_LIST:
            if option['id'] == int(employeeId):
                employeeName = option['name']
                break
  
        if employeeId == '0':
            error_message = "Select Employee"
            return render(request, settings.TEMPLATE_INDEX, {"employee_list": EMPLOYEE_LIST, 'error': error_message})
        else:
            print(request.POST)
            request.session['employeeId'] = employeeId
            request.session['employeeName'] = employeeName
            if 'Calendar.x' in request.POST:
                return redirect('calendar_view')
            elif 'Seats.x' in request.POST:
                return redirect('seat_view')
        
    return render(request, settings.TEMPLATE_INDEX, {"employee_list": EMPLOYEE_LIST})
