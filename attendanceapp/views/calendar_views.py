from django.shortcuts import render, redirect
from ..utils import get_target_month, get_all_sheets
from django.conf import settings
import openpyxl

def initialize(request):
    employeeId = request.session.get('employeeId')
    employeeName = request.session.get('employeeName')

    if request.session.get('selected_month') is not None:
        selected_month = request.session.get('selected_month')
        del request.session['selected_month']
    else:
        selected_month = get_target_month()

    errMessage = ''
    successMessage = ''
    if request.session.get('err') is not None:
        errMessage = request.session.get('err')
        del request.session['err']
    if request.session.get('success') is not None:
        successMessage = request.session.get('success')
        del request.session['success']

    headers, is_weekend_flags, slots, data = get_reservations(selected_month)

    return render(
        request,
        settings.TEMPLATE_CALENDAR,
        {
            "employeeId": employeeId,
            "employeeName": employeeName,
            "headers": headers,
            "is_weekend_flags": is_weekend_flags,
            "slots": slots,
            "data": data,
            "month_list": get_all_sheets(),
            "selected_month": selected_month,
            "errMessage": errMessage,
            "successMessage": successMessage
        },
    )

def set_reservation(employeeId, seat_id, colNo):
    wb = openpyxl.load_workbook(settings.EXCEL_PATH)

    isSuccess = False
    try:
        currentSheet = wb[get_target_month()]
        existValue = currentSheet.cell(row=int(seat_id) + 2, column= int(colNo) + 2).value
        if existValue is None:
            currentSheet.cell(row=int(seat_id) + 2, column= int(colNo) + 2, value=employeeId)
            wb.save(settings.EXCEL_PATH)
            isSuccess = True
    except:
        pass
    finally:
        wb.close()
        return isSuccess

def get_reservations(month):
    wb = openpyxl.load_workbook(settings.EXCEL_PATH)
    currentSheet = wb[month]
    headers = []
    slots = []
    data = []
    is_weekend_flags = []

    for row in currentSheet.iter_rows(min_row=1, max_row=currentSheet.max_row, values_only=True):
        if len(headers) == 0:
            for col in row[1:]:
                if col is None: continue
                headers.append(col)
                is_weekend_flags.append(col in ["Sat", "Sun"])
                
        else:
            if row[0] is None: continue
            if len(slots) == 0:
                slots.extend(row[1:])
            else:
                seat_id = row[0]
                am_pm_data = [value if value is not None else "" for value in row[1:]]
                data.append({"seat_id": seat_id, "am_pm_data":am_pm_data})

    return [headers, is_weekend_flags, slots, data]


def reserve_check(request):

    employeeId = request.session.get('employeeId')
    seat_id = request.POST.get("seat_id")
    column_number = request.POST.get("column_number")
    isSuccess = set_reservation(employeeId, seat_id, column_number)

    if isSuccess == False:
        request.session['err'] = 'Already taken'
    if isSuccess == True:
        request.session['success'] = 'Reservation success'

    return redirect('calendar_view') 

def cancel_reservation(request):
    seat_id = request.POST.get('seat_id')
    column_number = request.POST.get('column_number')

    wb = openpyxl.load_workbook(settings.EXCEL_PATH)
    try:
        currentSheet = wb[get_target_month()]
        currentSheet.cell(row=int(seat_id) + 2, column= int(column_number) + 2, value='')
        wb.save(settings.EXCEL_PATH)
        return redirect('calendar_view') 
    except:
        pass
    finally:
        wb.close()

    return redirect('calendar_view') 

def month_change(request):
    if request.method == 'POST':
        selected_month = request.POST.get('monthSelect')
        request.session['selected_month'] = selected_month

    return redirect('calendar_view') 