import openpyxl
from django.conf import settings
from datetime import datetime
import re


def get_target_month():
    return datetime.now().strftime("%Y-%m")

def get_employee_list():
    wb = openpyxl.load_workbook(settings.EXCEL_PATH)
    empSheet = wb[settings.EXCEL_SHEET_NAME_EMPLOYEE]
    employeeList = []
    for row in empSheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            employeeList.append({"id": row[0], "name": row[1]})

    return employeeList

def get_all_sheets():
    monthlySheets = []
    wb = openpyxl.load_workbook(settings.EXCEL_PATH)
    pattern = re.compile(r'^\d{4}-\d{2}$')
    monthlySheets = [sheet for sheet in wb.sheetnames if pattern.match(sheet)]
    return monthlySheets